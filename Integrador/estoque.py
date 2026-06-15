import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side #classes do openpyxl que representam cada aspecto visual de uma célula:
from openpyxl.utils import get_column_letter #função utilitária do openpyxl que converte um número de coluna na letra correspondente do Excel
from datetime import datetime

ARQUIVO_CSV_ENTRADA = "Produtos.csv"
ARQUIVO_DADOS_JS    = "dados.js"
ARQUIVO_SCRIPT_JS   = "script.js"
ARQUIVO_HTML        = "estoque.html"
ARQUIVO_CSS         = "estoque.css"
ARQUIVO_XLSX        = "relatorio_estoque.xlsx"
ARQUIVO_CSV_SAIDA   = "relatorio_estoque.csv"

LIMITE_CRITICO = 20
LIMITE_MEDIO   = 80

COR_HEADER  = "1A237E"
COR_CRITICO = "FFCCCC"
COR_MEDIO   = "FFF3CD"
COR_CHEIO   = "D4EDDA"
COR_ZEBRA   = "F5F5F5"

#LEITURA E PREPARAÇÃO DOS DADOS

def ler_estoque(caminho: str) -> pd.DataFrame:
    #Lê o CSV de estoque e retorna um DataFrame tratado
    try: #trata 3 tipos de erro: arquivo não encontrado, falta de permissão de leitura e falha na interpretação do csv
        df = pd.read_csv(caminho, sep=";", encoding="utf-8", dtype={"Codigo": str})
    except FileNotFoundError:
        raise FileNotFoundError(f"Arquivo CSV não encontrado: {caminho}")
    except PermissionError:
        raise PermissionError(f"Sem permissão para ler: {caminho}")
    except pd.errors.ParserError as erro:
        raise ValueError(f"Erro ao interpretar o CSV: {erro}")

    df["Preco"] = (
        df["Preco"].astype(str)
        .str.replace(".", "", regex=False)
        .str.replace(",", ".", regex=False)
        .astype(float)
    )

    print(f"[OK] {len(df)} produtos carregados do CSV.")
    return df


def classificar_nivel(qtde_livre: int) -> str:
    #Retorna o nível do estoque com base na quantidade livre
    if qtde_livre <= LIMITE_CRITICO:
        return "Crítico"
    if qtde_livre <= LIMITE_MEDIO:
        return "Médio"
    return "Cheio"


def calcular_colunas(df: pd.DataFrame) -> pd.DataFrame:
    #Adiciona colunas calculadas: quantidade livre, valor total e nível
    df = df.copy()
    df["Qtde_Livre"]  = df["Qtde_Disponivel"] - df["Qtde_Reservada"]
    df["Valor_Total"] = df["Qtde_Disponivel"] * df["Preco"]
    df["Nivel"] = df["Qtde_Livre"].apply(classificar_nivel) #aplica a função de classificação para cada linha, através do apply, ao invés de criar um loop para percorrer as linhas manualmente
    return df

#DADOS PARA A PÁGINA WEB (json)

def gerar_dados_js(df: pd.DataFrame, caminho: str) -> None:
    #Converte o DataFrame em JSON e escreve dados.js (variável global)
    registros = df.to_dict(orient="records")

    resumo = {
        "total_produtos": len(df),
        "criticos": int((df["Nivel"] == "Crítico").sum()),
        "medios":   int((df["Nivel"] == "Médio").sum()),
        "cheios":   int((df["Nivel"] == "Cheio").sum()),
        "atualizado_em": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
    }

    conteudo_json = json.dumps( #Converte um dicionário ou lista Python em uma string de texto no formato JSON
        {"resumo": resumo, "produtos": registros},
        ensure_ascii=False, #acentos e caracteres especiais são mantidos no arquivo
        indent=2 #formata o json de forma legível
    )

    try:
        with open(caminho, mode="w", encoding="utf-8") as arquivo:
            arquivo.write(f"const ESTOQUE_DATA = {conteudo_json};\n")
        print(f"[OK] Dados gerados: {caminho}")
    except PermissionError:
        raise PermissionError(f"Sem permissão para escrever: {caminho}")

#SCRIPT JS QUE MONTA A PÁGINA
def gerar_script_js(caminho: str) -> None:
    #Escreve script.js: monta cards e tabela a partir de ESTOQUE_DATA
    conteudo = """// script.js – monta a página de estoque a partir de ESTOQUE_DATA (dados.js)
#formata um número em padrão da moeda brasileira
function formatarMoeda(valor) {
  return valor.toLocaleString("pt-BR", { style: "currency", currency: "BRL" });
}

#cria um badge colorido para o nível do estoque
function badgeNivel(nivel) {
  const classes = {
    "Crítico": "nivel-critico",
    "Médio": "nivel-medio",
    "Cheio": "nivel-cheio",
  };
  return `<span class="badge ${classes[nivel] || ""}">${nivel}</span>`;
}

#monta os cards de resumo (total, críticos, médios, cheios) e coloca os coloca no html
function montarCards(resumo) {
  const container = document.getElementById("cards");
  container.innerHTML = `
    <div class="card">
      <div class="valor">${resumo.total_produtos}</div>
      <div class="label">Total de Produtos</div>
    </div>
    <div class="card">
      <div class="valor cor-critico">${resumo.criticos}</div>
      <div class="label">Estoque Crítico</div>
    </div>
    <div class="card">
      <div class="valor cor-medio">${resumo.medios}</div>
      <div class="label">Estoque Médio</div>
    </div>
    <div class="card">
      <div class="valor cor-cheio">${resumo.cheios}</div>
      <div class="label">Estoque Cheio</div>
    </div>
  `;
}

#monta as linhas da tabela de produtos e insere no corpo da tabela
function montarTabela(produtos) {
  const corpo = document.getElementById("corpo-tabela");
  corpo.innerHTML = produtos.map((p) => `
    <tr>
      <td>${p.Codigo}</td>
      <td class="nome">${p.Produto}</td>
      <td>${p.Qtde_Disponivel}</td>
      <td>${p.Qtde_Reservada}</td>
      <td><strong>${p.Qtde_Livre}</strong></td>
      <td>${formatarMoeda(p.Preco)}</td>
      <td>${formatarMoeda(p.Valor_Total)}</td>
      <td>${badgeNivel(p.Nivel)}</td>
    </tr>
  `).join("");
}

function montarRodape(resumo) {
  document.getElementById("data-atualizacao").textContent = resumo.atualizado_em;
}

function inicializar() {
  montarCards(ESTOQUE_DATA.resumo);
  montarTabela(ESTOQUE_DATA.produtos);
  montarRodape(ESTOQUE_DATA.resumo);
}

document.addEventListener("DOMContentLoaded", inicializar);
"""
    try: #se o arquivo já existir e estiver aberto em outro programa, ou o programa não tem permissão do sistema p/ criar arquivos na pasta
        with open(caminho, mode="w", encoding="utf-8") as arquivo:
            arquivo.write(conteudo)
        print(f"[OK] Script gerado: {caminho}")
    except PermissionError:
        raise PermissionError(f"Sem permissão para escrever: {caminho}")


#CSS - só para deixar a página mais bonitinha

def gerar_css(caminho: str) -> None:
    #Escreve estoque.css com o estilo da página
    conteudo = """* { box-sizing: border-box; margin: 0; padding: 0; }

body {
  font-family: Arial, sans-serif;
  background: #f0f2f5;
  color: #333;
}

header {
  background: #1a237e;
  color: #fff;
  padding: 20px 32px;
}

header h1 { font-size: 1.6em; }
header p  { font-size: .9em; opacity: .8; margin-top: 4px; }

.cards {
  display: flex;
  gap: 16px;
  padding: 24px 32px 0;
  flex-wrap: wrap;
}

.card {
  background: #fff;
  border-radius: 10px;
  padding: 16px 24px;
  box-shadow: 0 2px 6px rgba(0,0,0,.08);
  min-width: 160px;
}

.card .valor { font-size: 2em; font-weight: 700; color: #1a237e; }
.card .label { font-size: .85em; color: #666; margin-top: 4px; }

.cor-critico { color: #c0392b !important; }
.cor-medio   { color: #e67e22 !important; }
.cor-cheio   { color: #27ae60 !important; }

.container { padding: 24px 32px; }

table {
  width: 100%;
  border-collapse: collapse;
  background: #fff;
  border-radius: 10px;
  overflow: hidden;
  box-shadow: 0 2px 6px rgba(0,0,0,.08);
}

thead { background: #1a237e; color: #fff; }

th { padding: 12px 14px; text-align: left; font-size: .9em; }

td {
  padding: 10px 14px;
  font-size: .88em;
  border-bottom: 1px solid #eee;
  text-align: center;
}

td.nome { text-align: left; }

tbody tr:nth-child(even) { background: #f9f9f9; }
tbody tr:last-child td { border-bottom: none; }

.badge {
  padding: 2px 10px;
  border-radius: 12px;
  font-size: .85em;
  font-weight: 600;
}

.nivel-critico { background: #ffe0e0; color: #c0392b; }
.nivel-medio   { background: #fff8e1; color: #e67e22; }
.nivel-cheio   { background: #e8f5e9; color: #27ae60; }

footer {
  text-align: center;
  padding: 20px;
  font-size: .8em;
  color: #999;
}
"""
    #executa a mesma função do try anterior, mas para o arquivo CSS. Se o arquivo já existir e estiver aberto em outro programa, ou o programa não tem permissão do sistema p/ criar arquivos na pasta, ele vai lançar um erro de permissão.
    try:
        with open(caminho, mode="w", encoding="utf-8") as arquivo:
            arquivo.write(conteudo)
        print(f"[OK] CSS gerado: {caminho}")
    except PermissionError:
        raise PermissionError(f"Sem permissão para escrever: {caminho}")

#HTML (apenas estrutura, dados via JS)

def gerar_html(caminho: str, css_path: str, dados_js_path: str, script_js_path: str) -> None:
    #Escreve a página HTML, que carrega CSS e os arquivos JS de dados/script.
    conteudo = f"""<!DOCTYPE html>
<html lang="pt-BR">
<head>
  <meta charset="UTF-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1.0" />
  <title>Estoque – E-Commerce</title>
  <link rel="stylesheet" href="{css_path}" />
</head>
<body>
  <header>
    <h1> Painel de Estoque</h1>
    <p>Atualizado em: <span id="data-atualizacao">--</span></p>
  </header>

  <div class="cards" id="cards"></div>

  <div class="container">
    <table>
      <thead>
        <tr>
          <th>Código</th><th>Produto</th><th>Disponível</th>
          <th>Reservado</th><th>Livre</th><th>Preço Unit.</th>
          <th>Valor Total</th><th>Nível</th>
        </tr>
      </thead>
      <tbody id="corpo-tabela"></tbody>
    </table>
  </div>

  <footer>Gerado automaticamente pelo sistema de estoque</footer>

  <script src="{dados_js_path}"></script>
  <script src="{script_js_path}"></script>
</body>
</html>"""
    #executa a mesma função do try anterior, mas para o arquivo HTML. 
    try:
        with open(caminho, mode="w", encoding="utf-8") as arquivo:
            arquivo.write(conteudo)
        print(f"[OK] Página HTML gerada: {caminho}")
    except PermissionError:
        raise PermissionError(f"Sem permissão para escrever: {caminho}")

#RELATÓRIO EXCEL (openpyxl)

def _borda_fina() -> Border:
    lado = Side(style="thin", color="CCCCCC")
    return Border(left=lado, right=lado, top=lado, bottom=lado)

def _cor_nivel(nivel: str) -> str:
    mapa = {"Crítico": COR_CRITICO, "Médio": COR_MEDIO, "Cheio": COR_CHEIO}
    return mapa.get(nivel, "FFFFFF")

def _aplicar_cabecalho(ws, colunas: list, linha: int = 1) -> None:
    for col_idx, titulo in enumerate(colunas, start=1):
        cell = ws.cell(row=linha, column=col_idx, value=titulo)
        cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill("solid", start_color=COR_HEADER)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _borda_fina()
    ws.row_dimensions[linha].height = 22

def _criar_aba_estoque(wb: Workbook, df: pd.DataFrame) -> None:
    #Cria a aba 'Estoque' com os dados detalhados de cada produto
    ws = wb.active #acessa essa aba que já existe e a guarda na variável ws
    ws.title = "Estoque" #simplesmente renomeia essa aba

    #nomeia as colunas do excel, e depois chama a função de aplicar o cabeçalho, que formata a linha do cabeçalho (cor, fonte, alinhamento, borda)
    colunas = ["Código", "Produto", "Qtde Disponível", "Qtde Reservada",
               "Qtde Livre", "Preço Unit. (R$)", "Valor Total (R$)", "Nível"]
    _aplicar_cabecalho(ws, colunas)

    for i, row in enumerate(df.itertuples(index=False), start=2):
        valores = [row.Codigo, row.Produto, row.Qtde_Disponivel, row.Qtde_Reservada,
                   row.Qtde_Livre, row.Preco, row.Valor_Total, row.Nivel]

        cor_bg = _cor_nivel(row.Nivel)
        for col_idx, valor in enumerate(valores, start=1):
            cell = ws.cell(row=i, column=col_idx, value=valor)
            cell.font = Font(name="Arial", size=10)
            cell.fill = PatternFill("solid", start_color=cor_bg)
            cell.alignment = Alignment(horizontal="center" if col_idx != 2 else "left",
                                        vertical="center")
            cell.border = _borda_fina()

        ws.cell(i, 6).number_format = 'R$ #,##0.00'
        ws.cell(i, 7).number_format = 'R$ #,##0.00'

    ultima = len(df) + 2
    ws.cell(ultima, 2, "TOTAIS").font = Font(name="Arial", bold=True, size=10)
    ws.cell(ultima, 3, f"=SUM(C2:C{ultima-1})").font = Font(name="Arial", bold=True)
    ws.cell(ultima, 7, f"=SUM(G2:G{ultima-1})").font = Font(name="Arial", bold=True)
    ws.cell(ultima, 7).number_format = 'R$ #,##0.00'
    for col in range(1, 9):
        ws.cell(ultima, col).fill   = PatternFill("solid", start_color="E8EAF6")
        ws.cell(ultima, col).border = _borda_fina()

    larguras = [16, 48, 18, 18, 12, 18, 20, 12]
    for col_idx, largura in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = largura

    ws.freeze_panes = "A2"

def _criar_aba_resumo(wb: Workbook, df: pd.DataFrame) -> None:
    #Cria a aba 'Resumo' com indicadores gerais do estoque
    ws = wb.create_sheet("Resumo")
    agora = datetime.now().strftime("%d/%m/%Y %H:%M") #formatação da data

    #define as configurações dos textos
    ws["A1"] = "RESUMO DO ESTOQUE"
    ws["A1"].font = Font(name="Arial", bold=True, size=14, color="1A237E")
    ws["A2"] = f"Atualizado em: {agora}"
    ws["A2"].font = Font(name="Arial", italic=True, size=10, color="888888")

    _aplicar_cabecalho(ws, ["Indicador", "Valor"], linha=4)

    totais = {
        "Total de Produtos": len(df),
        f"Estoque Crítico (≤{LIMITE_CRITICO} livres)": int((df["Nivel"] == "Crítico").sum()),
        "Estoque Médio": int((df["Nivel"] == "Médio").sum()),
        "Estoque Cheio": int((df["Nivel"] == "Cheio").sum()),
    }

    for idx, (indicador, valor) in enumerate(totais.items(), start=5):
        cor = COR_ZEBRA if idx % 2 == 0 else "FFFFFF"
        ws.cell(idx, 1, indicador).font = Font(name="Arial", size=10)
        ws.cell(idx, 2, valor).font     = Font(name="Arial", size=10, bold=True)
        for col in [1, 2]:
            ws.cell(idx, col).fill      = PatternFill("solid", start_color=cor)
            ws.cell(idx, col).border    = _borda_fina()
            ws.cell(idx, col).alignment = Alignment(horizontal="center" if col == 2 else "left")

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 14

def gerar_excel(df: pd.DataFrame, caminho: str) -> None:
    #Cria o relatório Excel com abas 'Estoque' e 'Resumo'.
    wb = Workbook()
    _criar_aba_estoque(wb, df)
    _criar_aba_resumo(wb, df)

    try:
        wb.save(caminho)
        print(f"[OK] Relatório Excel gerado: {caminho}")
    except PermissionError:
        raise PermissionError(f"Sem permissão para escrever: {caminho}") #mesma coisa, se o arquivo já estiver aberto

#GERAÇÃO DO RELATÓRIO CSV

def gerar_csv(df: pd.DataFrame, caminho: str) -> None:
    #Exporta o DataFrame final (com colunas calculadas) para CSV.
    colunas = ["Codigo", "Produto", "Qtde_Disponivel", "Qtde_Reservada",
               "Qtde_Livre", "Preco", "Valor_Total", "Nivel"]

    try:
        df[colunas].to_csv(caminho, sep=";", index=False, encoding="utf-8",
                           decimal=",") #mantém o padrão brasileiro de separação decimal e de campo, e não inclui o índice do DataFrame
        print(f"[OK] Relatório CSV gerado: {caminho}")
    except PermissionError:
        raise PermissionError(f"Sem permissão para escrever: {caminho}")
    except OSError as erro:
        raise OSError(f"Erro ao salvar CSV: {erro}")

#ORQUESTRAÇÃO PRINCIPAL

def processar_estoque() -> None:
    #Executa todo o fluxo de automação logística
    print("=" * 52)
    print("  SISTEMA DE AUTOMAÇÃO LOGÍSTICA")
    print("=" * 52)

    try:
        df = ler_estoque(ARQUIVO_CSV_ENTRADA)
        df = calcular_colunas(df)

        gerar_dados_js(df, ARQUIVO_DADOS_JS)
        gerar_script_js(ARQUIVO_SCRIPT_JS)
        gerar_css(ARQUIVO_CSS)
        gerar_html(ARQUIVO_HTML, ARQUIVO_CSS, ARQUIVO_DADOS_JS, ARQUIVO_SCRIPT_JS)

        gerar_excel(df, ARQUIVO_XLSX)
        gerar_csv(df, ARQUIVO_CSV_SAIDA)

        print("=" * 52)
        print("  PROCESSAMENTO CONCLUÍDO COM SUCESSO!")
        print("=" * 52)

    except FileNotFoundError as erro:
        print(f"[ERRO] Arquivo não encontrado: {erro}")
    except PermissionError as erro:
        print(f"[ERRO] Permissão negada: {erro}")
    except Exception as erro:
        print(f"[ERRO] Falha inesperada: {erro}")
        raise
'''Esse bloco foi criado para garantir que a função processar_estoque só seja executada 
quando o arquivo estoque.py for rodado diretamente pelo terminal.'''
if __name__ == "__main__":
    processar_estoque()
